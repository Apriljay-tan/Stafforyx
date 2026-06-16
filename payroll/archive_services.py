"""
Payroll Archive & Cleanup services.

Pure, view-independent helpers that:
  * collect the payroll-related records for a company + date range,
  * count them (for the preview step),
  * build a multi-sheet .xlsx workbook (export step), and
  * delete only those records inside a transaction (cleanup step).

Hard safety rules enforced here:
  * Export NEVER deletes anything.
  * Cleanup only ever touches the record types listed in ``ARCHIVE_MODELS``.
    Employees, companies, users, company/payroll settings, attendance
    locations, kiosk devices and payroll periods are never deleted.
  * If an optional app/model is not installed, it is skipped gracefully.
"""

import os
import re
from datetime import datetime, date, time
from decimal import Decimal

from django.conf import settings
from django.db import transaction
from django.utils import timezone

from .models import PayrollRecord, PayrollPeriod


ARCHIVE_SUBDIR = 'payroll_archives'

# Logical archive sections in a fixed, FK-safe deletion order. Logs and
# requests are deleted before the attendance/payroll rows they may reference
# (those references are SET_NULL, so order is defensive rather than required).
SECTION_ORDER = [
    'portal_logs',
    'qr_logs',
    'overtime_requests',
    'leave_requests',
    'ca_requests',
    'attendance_records',
    'payroll_records',
]


# ── Optional-model resolution ───────────────────────────────────────────────────

def _safe_import(path, name):
    """Return a model class or None if its app/model is unavailable."""
    try:
        module = __import__(path, fromlist=[name])
        return getattr(module, name, None)
    except Exception:
        return None


AttendanceRecord = _safe_import('attendance.models', 'AttendanceRecord')
AttendancePortalLog = _safe_import('attendance.models', 'AttendancePortalLog')
AttendanceQRScanLog = _safe_import('attendance.models', 'AttendanceQRScanLog')
OvertimeRequest = _safe_import('overtime.models', 'OvertimeRequest')
LeaveRequest = _safe_import('leaves.models', 'LeaveRequest')
CashAdvanceRequest = _safe_import('cash_advance.models', 'CashAdvanceRequest')


# ── Date-range resolution ───────────────────────────────────────────────────────

def resolve_date_range(date_from, date_to, payroll_period=None):
    """A selected payroll period defines its own start/end date range."""
    if payroll_period is not None:
        return payroll_period.start_date, payroll_period.end_date
    return date_from, date_to


# ── Record collection ───────────────────────────────────────────────────────────

def collect_archive_querysets(company, date_from, date_to, payroll_period=None):
    """
    Return a dict of {section: queryset} scoped to one company and the given
    date range. Missing optional models yield ``None`` for that section.
    """
    querysets = {}

    if payroll_period is not None:
        payroll_qs = PayrollRecord.objects.filter(
            company=company, payroll_period=payroll_period
        )
    else:
        # Only fully-contained periods are archived, so an in-progress period
        # that merely overlaps the range is never swept up.
        payroll_qs = PayrollRecord.objects.filter(
            company=company,
            payroll_period__start_date__gte=date_from,
            payroll_period__end_date__lte=date_to,
        )
    querysets['payroll_records'] = payroll_qs.select_related(
        'employee', 'payroll_period', 'company'
    )

    if AttendanceRecord is not None:
        querysets['attendance_records'] = AttendanceRecord.objects.filter(
            company=company, date__gte=date_from, date__lte=date_to,
        ).select_related('employee', 'portal_location')
    else:
        querysets['attendance_records'] = None

    if AttendancePortalLog is not None:
        querysets['portal_logs'] = AttendancePortalLog.objects.filter(
            company=company,
            created_at__date__gte=date_from, created_at__date__lte=date_to,
        ).select_related('employee', 'attendance_location')
    else:
        querysets['portal_logs'] = None

    if AttendanceQRScanLog is not None:
        querysets['qr_logs'] = AttendanceQRScanLog.objects.filter(
            company=company,
            created_at__date__gte=date_from, created_at__date__lte=date_to,
        ).select_related('employee', 'attendance_location', 'kiosk_device')
    else:
        querysets['qr_logs'] = None

    if OvertimeRequest is not None:
        querysets['overtime_requests'] = OvertimeRequest.objects.filter(
            company=company, date__gte=date_from, date__lte=date_to,
        ).select_related('employee', 'reviewed_by')
    else:
        querysets['overtime_requests'] = None

    if LeaveRequest is not None:
        # Leaves overlapping the range (start before range end, end after range start).
        querysets['leave_requests'] = LeaveRequest.objects.filter(
            company=company, start_date__lte=date_to, end_date__gte=date_from,
        ).select_related('employee', 'leave_type', 'reviewed_by')
    else:
        querysets['leave_requests'] = None

    if CashAdvanceRequest is not None:
        querysets['ca_requests'] = CashAdvanceRequest.objects.filter(
            company=company,
            created_at__date__gte=date_from, created_at__date__lte=date_to,
        ).select_related('employee', 'approved_by', 'released_by')
    else:
        querysets['ca_requests'] = None

    return querysets


def count_archive_records(querysets):
    """Return {section: count}; missing sections report 0."""
    return {
        key: (qs.count() if qs is not None else 0)
        for key, qs in querysets.items()
    }


# ── Value normalisation for openpyxl ────────────────────────────────────────────

def _clean(value):
    """Coerce a model value into something openpyxl can write safely."""
    if value is None:
        return ''
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime):
        # openpyxl cannot serialise tz-aware datetimes — store naive local time.
        if timezone.is_aware(value):
            value = timezone.localtime(value)
            return value.replace(tzinfo=None)
        return value
    if isinstance(value, time):
        return value.strftime('%H:%M:%S')
    if isinstance(value, date):
        return value
    return value


def _emp_name(emp):
    if emp is None:
        return ''
    return getattr(emp, 'full_name', None) or str(emp)


def _emp_code(emp):
    return getattr(emp, 'employee_id', '') if emp is not None else ''


# ── Sheet definitions: (section, title, headers, row(obj) -> list) ──────────────

def _payroll_row(r):
    return [
        _emp_code(r.employee), _emp_name(r.employee),
        r.payroll_period.name if r.payroll_period_id else '',
        r.company.name if r.company_id else '',
        _clean(r.basic_pay), _clean(r.gross_pay),
        _clean(getattr(r, 'total_deductions', None)) if hasattr(r, 'total_deductions') else '',
        _clean(r.net_pay), r.get_status_display(),
        _clean(getattr(r, 'overtime_minutes', None)), _clean(getattr(r, 'overtime_pay', None)),
        _clean(getattr(r, 'overtime_multiplier', None)),
        _clean(getattr(r, 'night_differential_minutes', None)),
        _clean(getattr(r, 'night_differential_pay', None)),
        _clean(getattr(r, 'holiday_pay', None)),
        _clean(r.created_at), _clean(r.updated_at),
    ]


_PAYROLL_HEADERS = [
    'Employee ID', 'Employee', 'Payroll Period', 'Company',
    'Basic Pay', 'Gross Pay', 'Deductions', 'Net Pay', 'Status',
    'OT Minutes', 'OT Pay', 'OT Multiplier',
    'Night Diff Minutes', 'Night Diff Pay', 'Holiday Pay',
    'Created At', 'Updated At',
]


def _attendance_row(r):
    return [
        _emp_code(r.employee), _emp_name(r.employee),
        r.company.name if r.company_id else '',
        r.portal_location.name if getattr(r, 'portal_location_id', None) else '',
        _clean(r.date), _clean(r.time_in), _clean(r.time_out),
        _clean(getattr(r, 'total_hours', None)),
        _clean(getattr(r, 'total_work_minutes', None)),
        _clean(getattr(r, 'overtime_minutes', None)),
        _clean(getattr(r, 'night_differential_minutes', None)),
        r.get_status_display() if hasattr(r, 'get_status_display') else _clean(getattr(r, 'status', None)),
        _clean(getattr(r, 'computed_status', None)),
        _clean(getattr(r, 'source', None)),
        _clean(getattr(r, 'remarks', None)),
    ]


_ATTENDANCE_HEADERS = [
    'Employee ID', 'Employee', 'Company', 'Location',
    'Date', 'Time In', 'Time Out', 'Total Hours', 'Worked Minutes',
    'OT Minutes', 'Night Diff Minutes', 'Status', 'Computed Status',
    'Source', 'Remarks',
]


def _portal_log_row(r):
    return [
        _emp_name(r.employee),
        r.company.name if r.company_id else '',
        r.attendance_location.name if getattr(r, 'attendance_location_id', None) else '',
        _clean(getattr(r, 'action', None)),
        _clean(getattr(r, 'status', None)),
        _clean(getattr(r, 'validation_method', None)),
        _clean(r.created_at),
        _clean(getattr(r, 'ip_address', None)),
        _clean(getattr(r, 'gps_latitude', None)),
        _clean(getattr(r, 'gps_longitude', None)),
        _clean(getattr(r, 'gps_accuracy', None)),
        _clean(getattr(r, 'blocked_reason', None)),
    ]


_PORTAL_LOG_HEADERS = [
    'Employee', 'Company', 'Location', 'Action', 'Status', 'Validation Method',
    'Timestamp', 'IP Address', 'GPS Lat', 'GPS Lng', 'GPS Accuracy', 'Blocked Reason',
]


def _qr_log_row(r):
    return [
        _emp_name(r.employee),
        r.company.name if r.company_id else '',
        r.attendance_location.name if getattr(r, 'attendance_location_id', None) else '',
        r.kiosk_device.name if getattr(r, 'kiosk_device_id', None) else '',
        _clean(getattr(r, 'action', None)),
        _clean(getattr(r, 'result', None)),
        _clean(r.created_at),
        _clean(getattr(r, 'ip_address', None)),
        _clean(getattr(r, 'gps_latitude', None)),
        _clean(getattr(r, 'gps_longitude', None)),
        _clean(getattr(r, 'token_hash', None)),
    ]


_QR_LOG_HEADERS = [
    'Employee', 'Company', 'Location', 'Kiosk Device', 'Action', 'Result',
    'Timestamp', 'IP Address', 'GPS Lat', 'GPS Lng', 'Token Hash',
]


def _overtime_row(r):
    return [
        _emp_code(r.employee), _emp_name(r.employee),
        _clean(r.date),
        _clean(getattr(r, 'requested_hours', None)),
        _clean(getattr(r, 'approved_hours', None)),
        r.get_status_display() if hasattr(r, 'get_status_display') else _clean(getattr(r, 'status', None)),
        getattr(r.reviewed_by, 'username', '') if getattr(r, 'reviewed_by_id', None) else '',
        _clean(getattr(r, 'reason', None)),
    ]


_OVERTIME_HEADERS = [
    'Employee ID', 'Employee', 'Date', 'Requested Hours', 'Approved Hours',
    'Status', 'Reviewed By', 'Reason',
]


def _leave_row(r):
    return [
        _emp_code(r.employee), _emp_name(r.employee),
        r.leave_type.name if getattr(r, 'leave_type_id', None) else '',
        _clean(r.start_date), _clean(r.end_date),
        _clean(getattr(r, 'total_days', None)),
        r.get_status_display() if hasattr(r, 'get_status_display') else _clean(getattr(r, 'status', None)),
        getattr(r.reviewed_by, 'username', '') if getattr(r, 'reviewed_by_id', None) else '',
        _clean(getattr(r, 'reason', None)),
    ]


_LEAVE_HEADERS = [
    'Employee ID', 'Employee', 'Leave Type', 'Start Date', 'End Date',
    'Total Days', 'Status', 'Reviewed By', 'Reason',
]


def _ca_row(r):
    return [
        _emp_code(r.employee), _emp_name(r.employee),
        _clean(getattr(r, 'amount', None)),
        _clean(getattr(r, 'total_deducted_amount', None)),
        r.get_status_display() if hasattr(r, 'get_status_display') else _clean(getattr(r, 'status', None)),
        r.get_deduction_status_display() if hasattr(r, 'get_deduction_status_display') else '',
        getattr(r.released_by, 'username', '') if getattr(r, 'released_by_id', None) else '',
        _clean(getattr(r, 'released_at', None)),
        _clean(r.created_at),
    ]


_CA_HEADERS = [
    'Employee ID', 'Employee', 'Amount', 'Deducted Amount', 'Status',
    'Deduction Status', 'Released By', 'Released At', 'Created At',
]


_SHEETS = [
    ('payroll_records', 'Payroll Records', _PAYROLL_HEADERS, _payroll_row),
    ('attendance_records', 'Attendance', _ATTENDANCE_HEADERS, _attendance_row),
    ('portal_logs', 'Portal Logs', _PORTAL_LOG_HEADERS, _portal_log_row),
    ('qr_logs', 'QR Scan Logs', _QR_LOG_HEADERS, _qr_log_row),
    ('overtime_requests', 'Overtime', _OVERTIME_HEADERS, _overtime_row),
    ('leave_requests', 'Leaves', _LEAVE_HEADERS, _leave_row),
    ('ca_requests', 'Cash Advances', _CA_HEADERS, _ca_row),
]


# ── Workbook construction ───────────────────────────────────────────────────────

# ── Styling constants (Stafforyx brand: navy #0D1B2A, blue #1565C0) ─────────────

FONT_NAME = 'Calibri'
NAVY = '0D1B2A'
BLUE = '1565C0'
ALT_ROW = 'F2F6FC'
BORDER_GREY = 'D9D9D9'

PESO_FMT = '₱#,##0.00'      # ₱#,##0.00
DATE_FMT = 'yyyy-mm-dd'
DATETIME_FMT = 'yyyy-mm-dd hh:mm'
INT_FMT = '#,##0'
DEC_FMT = '#,##0.00'

_CURRENCY_HEADERS = {
    'Basic Pay', 'Gross Pay', 'Deductions', 'Net Pay', 'OT Pay',
    'Night Diff Pay', 'Holiday Pay', 'Amount', 'Deducted Amount',
}
_DATE_HEADERS = {'Date', 'Start Date', 'End Date'}
_DECIMAL_HEADERS = {
    'Total Hours', 'OT Multiplier', 'Requested Hours', 'Approved Hours', 'Total Days',
}
_INT_HEADERS = {'OT Minutes', 'Worked Minutes', 'Night Diff Minutes'}
_WRAP_HEADERS = {
    'Reason', 'Blocked Reason', 'Remarks', 'User Agent', 'Token Hash', 'Notes',
}


def _column_spec(header):
    """Return (number_format, horizontal_align, wrap) for a column header."""
    if header in _CURRENCY_HEADERS:
        return PESO_FMT, 'right', False
    if header in _DATE_HEADERS:
        return DATE_FMT, 'center', False
    if header.endswith('At') or header == 'Timestamp':
        return DATETIME_FMT, 'center', False
    if header in _DECIMAL_HEADERS:
        return DEC_FMT, 'right', False
    if header in _INT_HEADERS:
        return INT_FMT, 'right', False
    if header in _WRAP_HEADERS:
        return None, 'left', True
    return None, 'left', False


def _autosize(ws, headers):
    from openpyxl.utils import get_column_letter
    for col_idx, header in enumerate(headers, start=1):
        _, _, wrap = _column_spec(header)
        max_len = len(str(header))
        for row_idx in range(2, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        letter = get_column_letter(col_idx)
        if wrap:
            ws.column_dimensions[letter].width = min(max(max_len, 20), 42)
        else:
            ws.column_dimensions[letter].width = min(max(max_len + 2, 11), 46)


def _apply_data_sheet_style(ws, headers):
    """Brand the header row, band data rows, add borders/filters/formats."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill('solid', fgColor=NAVY)
    header_font = Font(name=FONT_NAME, bold=True, color='FFFFFF', size=11)
    base_font = Font(name=FONT_NAME, size=10)
    status_font = Font(name=FONT_NAME, size=10, bold=True, color=BLUE)
    alt_fill = PatternFill('solid', fgColor=ALT_ROW)
    side = Side(style='thin', color=BORDER_GREY)
    border = Border(left=side, right=side, top=side, bottom=side)

    n_cols = len(headers)
    specs = [_column_spec(h) for h in headers]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 26

    for row_idx in range(2, ws.max_row + 1):
        banded = (row_idx % 2 == 0)
        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            num_fmt, align, wrap = specs[col_idx - 1]
            cell.font = status_font if headers[col_idx - 1] == 'Status' else base_font
            cell.border = border
            cell.alignment = Alignment(horizontal=align, vertical='top', wrap_text=wrap)
            if num_fmt and (isinstance(cell.value, (int, float)) or hasattr(cell.value, 'year')):
                cell.number_format = num_fmt
            if banded:
                cell.fill = alt_fill

    ws.freeze_panes = 'A2'
    last_col = get_column_letter(n_cols)
    ws.auto_filter.ref = f'A1:{last_col}{max(ws.max_row, 1)}'
    ws.sheet_view.zoomScale = 100
    _autosize(ws, headers)


def _build_summary_sheet(ws, *, company, date_from, date_to, period_label,
                         generated_by, counts, total_gross, total_net):
    """Render the Summary sheet as a clean, report-style dashboard."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    navy = PatternFill('solid', fgColor=NAVY)
    blue = PatternFill('solid', fgColor=BLUE)
    label_font = Font(name=FONT_NAME, bold=True, size=10, color=NAVY)
    value_font = Font(name=FONT_NAME, size=10)
    section_font = Font(name=FONT_NAME, bold=True, size=11, color='FFFFFF')
    underline = Border(bottom=Side(style='thin', color='E6E6E6'))

    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 46

    # Title band
    ws.merge_cells('A1:B1')
    title = ws['A1']
    title.value = 'STAFFORYX DATA ARCHIVE REPORT'
    title.fill = navy
    title.font = Font(name=FONT_NAME, bold=True, size=18, color='FFFFFF')
    title.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 38

    ws.merge_cells('A2:B2')
    subtitle = ws['A2']
    subtitle.value = company.name
    subtitle.font = Font(name=FONT_NAME, bold=True, size=12, color=BLUE)
    subtitle.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 22

    state = {'r': 4}

    def section(name):
        row = state['r']
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        cell = ws.cell(row=row, column=1, value=name)
        cell.fill = blue
        cell.font = section_font
        cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[row].height = 22
        state['r'] += 1

    def kv(label, value, *, number_format=None, align='left'):
        row = state['r']
        lcell = ws.cell(row=row, column=1, value=label)
        lcell.font = label_font
        lcell.alignment = Alignment(horizontal='left', vertical='center')
        lcell.border = underline
        vcell = ws.cell(row=row, column=2, value=value)
        vcell.font = value_font
        vcell.alignment = Alignment(horizontal=align, vertical='center')
        vcell.border = underline
        if number_format:
            vcell.number_format = number_format
        state['r'] += 1

    def spacer():
        state['r'] += 1

    section('Report Information')
    kv('Company', company.name)
    if period_label:
        kv('Payroll Period', period_label)
    kv('Date From', date_from, number_format=DATE_FMT)
    kv('Date To', date_to, number_format=DATE_FMT)
    kv('Generated By', getattr(generated_by, 'username', '') if generated_by else 'System')
    kv('Generated At', _clean(timezone.now()), number_format=DATETIME_FMT)
    spacer()

    section('Record Counts')
    kv('Payroll Records', counts.get('payroll_records', 0), align='right')
    kv('Attendance Records', counts.get('attendance_records', 0), align='right')
    kv('Portal Logs', counts.get('portal_logs', 0), align='right')
    kv('QR Scan Logs', counts.get('qr_logs', 0), align='right')
    kv('Overtime Records', counts.get('overtime_requests', 0), align='right')
    kv('Leave Records', counts.get('leave_requests', 0), align='right')
    kv('Cash Advance Records', counts.get('ca_requests', 0), align='right')
    spacer()

    section('Financial Summary')
    kv('Total Gross Pay', float(total_gross), number_format=PESO_FMT, align='right')
    kv('Total Net Pay', float(total_net), number_format=PESO_FMT, align='right')


def build_archive_workbook(querysets, *, company, date_from, date_to,
                           generated_by=None, counts=None):
    """Build and return a styled openpyxl Workbook (summary + one sheet per section)."""
    from openpyxl import Workbook

    if counts is None:
        counts = count_archive_records(querysets)

    wb = Workbook()
    wb.properties.title = 'Stafforyx Data Archive Report'
    wb.properties.creator = 'Stafforyx HR'

    # Totals + a single-period label (presentation only — no query changes).
    payroll_qs = querysets.get('payroll_records')
    total_gross = total_net = Decimal('0.00')
    period_names = set()
    if payroll_qs is not None:
        for rec in payroll_qs:
            total_gross += (rec.gross_pay or Decimal('0.00'))
            total_net += (rec.net_pay or Decimal('0.00'))
            if rec.payroll_period_id:
                period_names.add(rec.payroll_period.name)
    period_label = period_names.pop() if len(period_names) == 1 else ''

    summary = wb.active
    summary.title = 'Summary'
    _build_summary_sheet(
        summary, company=company, date_from=date_from, date_to=date_to,
        period_label=period_label, generated_by=generated_by, counts=counts,
        total_gross=total_gross, total_net=total_net,
    )

    for section, title, headers, row_fn in _SHEETS:
        ws = wb.create_sheet(title=title)
        ws.append(headers)
        qs = querysets.get(section)
        if qs is not None:
            for obj in qs:
                try:
                    ws.append(row_fn(obj))
                except Exception:
                    # A single malformed row must never abort the whole export.
                    ws.append(['(row export error)'])
        _apply_data_sheet_style(ws, headers)

    return wb


def _slugify_company(name):
    slug = re.sub(r'[^A-Za-z0-9]+', '_', name or '').strip('_').upper()
    return slug or 'COMPANY'


def build_archive_filename(company, date_from, date_to):
    return (
        f'stafforyx_archive_{_slugify_company(company.name)}'
        f'_{date_from}_to_{date_to}.xlsx'
    )


def save_workbook(wb, company, date_from, date_to):
    """
    Persist the workbook under MEDIA_ROOT/payroll_archives/ and return
    (file_name, relative_path, absolute_path). A timestamp suffix guarantees a
    unique filename if the same range is exported more than once.
    """
    base_name = build_archive_filename(company, date_from, date_to)
    archive_dir = os.path.join(settings.MEDIA_ROOT, ARCHIVE_SUBDIR)
    os.makedirs(archive_dir, exist_ok=True)

    file_name = base_name
    abs_path = os.path.join(archive_dir, file_name)
    if os.path.exists(abs_path):
        stamp = timezone.now().strftime('%Y%m%d%H%M%S')
        root, ext = os.path.splitext(base_name)
        file_name = f'{root}_{stamp}{ext}'
        abs_path = os.path.join(archive_dir, file_name)

    wb.save(abs_path)
    rel_path = os.path.join(ARCHIVE_SUBDIR, file_name).replace('\\', '/')
    return file_name, rel_path, abs_path


# ── Cleanup ─────────────────────────────────────────────────────────────────────

def perform_cleanup(batch):
    """
    Delete only the records covered by an ArchiveBatch, inside a transaction.

    Returns a dict of {section: deleted_count}. The batch's own fields are NOT
    modified here — the caller marks it cleared after a successful return.
    """
    querysets = collect_archive_querysets(
        batch.company, batch.date_from, batch.date_to, batch.payroll_period
    )
    cleared = {}
    with transaction.atomic():
        for section in SECTION_ORDER:
            qs = querysets.get(section)
            if qs is None:
                continue
            model_label = qs.model._meta.label
            deleted_total, breakdown = qs.delete()
            cleared[section] = breakdown.get(model_label, 0) or 0
    return cleared
