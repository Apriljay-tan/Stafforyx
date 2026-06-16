import datetime
import tempfile
from decimal import Decimal
from io import StringIO
from unittest.mock import MagicMock, patch

from django.contrib.auth.models import User
from django.core.management import call_command
from django.test import TestCase, TransactionTestCase, override_settings
from django.urls import reverse
from django.utils import timezone

from accounts.models import UserCompanyAccess, UserProfile
from attendance.models import (
    AttendanceLocation, AttendancePortalLog, AttendanceQRScanLog, AttendanceRecord,
)
from cash_advance.models import CashAdvanceRequest
from companies.models import Company
from employees.models import Employee
from leaves.models import LeaveRequest, LeaveType
from overtime.models import OvertimeRequest
from payroll.archive_services import collect_archive_querysets, count_archive_records
from payroll.models import ArchiveBatch, PayrollPeriod, PayrollRecord

_FROM = datetime.date(2026, 6, 1)
_TO = datetime.date(2026, 6, 15)
_IN_RANGE = datetime.date(2026, 6, 10)
_OUT_RANGE = datetime.date(2026, 7, 10)
_MID_DT = timezone.make_aware(datetime.datetime(2026, 6, 10, 9, 0))
_OUT_DT = timezone.make_aware(datetime.datetime(2026, 7, 10, 9, 0))

_TMP_MEDIA = tempfile.mkdtemp()


def _emp(company, code):
    return Employee.objects.create(
        company=company, employee_id=code, first_name=code, last_name='W',
        email=f'{code.lower()}@t.com', date_hired=datetime.date(2020, 1, 1),
        basic_salary=Decimal('26000.00'), status='active',
    )


@override_settings(MEDIA_ROOT=_TMP_MEDIA)
@patch('licenses.clock_guard.check_clock_rollback', return_value=False)
@patch('licenses.middleware.is_license_active', return_value=True)
class ArchiveCenterTests(TestCase):
    def setUp(self):
        self.company = Company.objects.create(name='MB San Isidro', email='mb@test.com')
        self.other = Company.objects.create(name='Other Co', email='other@test.com')
        self.admin = User.objects.create_superuser('arch_admin', password='pass')

        self.period = PayrollPeriod.objects.create(
            company=self.company, name='June 1-15',
            start_date=_FROM, end_date=_TO,
        )
        self.emp = _emp(self.company, 'E1')
        self.other_emp = _emp(self.other, 'O1')
        self.location = AttendanceLocation.objects.create(
            company=self.company, name='HQ', ip_address='127.0.0.1',
        )

        # ── In-range records (company) ──
        self.payroll = PayrollRecord.objects.create(
            company=self.company, payroll_period=self.period, employee=self.emp,
            status='approved', basic_pay=Decimal('1000'), gross_pay=Decimal('1200'),
            net_pay=Decimal('1000'), daily_rate=Decimal('1000'), hourly_rate=Decimal('125'),
        )
        self.att = AttendanceRecord.objects.create(
            company=self.company, employee=self.emp, date=_IN_RANGE,
            time_in=datetime.time(8, 0), status='present',
        )
        self.portal_log = AttendancePortalLog.objects.create(
            company=self.company, employee=self.emp, action='time_in', status='success',
        )
        AttendancePortalLog.objects.filter(pk=self.portal_log.pk).update(created_at=_MID_DT)
        self.qr_log = AttendanceQRScanLog.objects.create(
            company=self.company, employee=self.emp, action='validate', result='success',
        )
        AttendanceQRScanLog.objects.filter(pk=self.qr_log.pk).update(created_at=_MID_DT)
        self.ot = OvertimeRequest.objects.create(
            company=self.company, employee=self.emp, date=_IN_RANGE,
            requested_hours=Decimal('2.0'), status='approved',
        )
        self.ltype = LeaveType.objects.create(company=self.company, name='VL')
        self.leave = LeaveRequest.objects.create(
            company=self.company, employee=self.emp, leave_type=self.ltype,
            start_date=_IN_RANGE, end_date=_IN_RANGE, total_days=1, reason='x', status='approved',
        )
        self.ca = CashAdvanceRequest.objects.create(
            company=self.company, employee=self.emp, amount=Decimal('500'),
        )
        CashAdvanceRequest.objects.filter(pk=self.ca.pk).update(created_at=_MID_DT)

        # ── Out-of-range / other-company records (must survive) ──
        self.att_out = AttendanceRecord.objects.create(
            company=self.company, employee=self.emp, date=_OUT_RANGE,
            time_in=datetime.time(8, 0), status='present',
        )
        self.other_att = AttendanceRecord.objects.create(
            company=self.other, employee=self.other_emp, date=_IN_RANGE,
            time_in=datetime.time(8, 0), status='present',
        )

        self.url = reverse('payroll:archive_center')

    def _login_admin(self):
        self.client.login(username='arch_admin', password='pass')

    # ── Preview ──────────────────────────────────────────────────────────────

    def test_preview_counts_only_in_range_company_records(self, *_m):
        self._login_admin()
        response = self.client.post(self.url, {
            'action': 'preview', 'company': self.company.pk,
            'date_from': _FROM, 'date_to': _TO,
        })
        self.assertEqual(response.status_code, 200)
        counts = response.context['preview']['counts']
        self.assertEqual(counts['payroll_records'], 1)
        self.assertEqual(counts['attendance_records'], 1)   # out-of-range excluded
        self.assertEqual(counts['portal_logs'], 1)
        self.assertEqual(counts['qr_logs'], 1)
        self.assertEqual(counts['overtime_requests'], 1)
        self.assertEqual(counts['leave_requests'], 1)
        self.assertEqual(counts['ca_requests'], 1)

    def test_preview_does_not_delete(self, *_m):
        self._login_admin()
        self.client.post(self.url, {
            'action': 'preview', 'company': self.company.pk,
            'date_from': _FROM, 'date_to': _TO,
        })
        self.assertTrue(PayrollRecord.objects.filter(pk=self.payroll.pk).exists())
        self.assertEqual(ArchiveBatch.objects.count(), 0)

    # ── Export ───────────────────────────────────────────────────────────────

    def test_export_creates_batch_and_file_without_deleting(self, *_m):
        import os
        from django.conf import settings
        from openpyxl import load_workbook

        self._login_admin()
        response = self.client.post(self.url, {
            'action': 'export', 'company': self.company.pk,
            'date_from': _FROM, 'date_to': _TO,
        })
        self.assertEqual(response.status_code, 302)  # redirects to download
        batch = ArchiveBatch.objects.get()
        self.assertEqual(batch.payroll_count, 1)
        self.assertFalse(batch.is_cleared)

        # Nothing deleted by export.
        self.assertTrue(PayrollRecord.objects.filter(pk=self.payroll.pk).exists())
        self.assertTrue(AttendanceRecord.objects.filter(pk=self.att.pk).exists())

        # File exists with the expected sheets.
        path = os.path.join(settings.MEDIA_ROOT, batch.file_path)
        self.assertTrue(os.path.exists(path))
        wb = load_workbook(path)
        for sheet in ['Summary', 'Payroll Records', 'Attendance', 'Portal Logs',
                      'QR Scan Logs', 'Overtime', 'Leaves', 'Cash Advances']:
            self.assertIn(sheet, wb.sheetnames)

    def test_filename_uses_company_slug(self, *_m):
        self._login_admin()
        self.client.post(self.url, {
            'action': 'export', 'company': self.company.pk,
            'date_from': _FROM, 'date_to': _TO,
        })
        batch = ArchiveBatch.objects.get()
        self.assertIn('MB_SAN_ISIDRO', batch.file_name)
        self.assertTrue(batch.file_name.endswith('.xlsx'))

    # ── Clear ────────────────────────────────────────────────────────────────

    def _make_batch(self):
        return ArchiveBatch.objects.create(
            company=self.company, date_from=_FROM, date_to=_TO,
            file_name='x.xlsx', file_path='payroll_archives/x.xlsx',
            generated_by=self.admin,
        )

    def test_clear_blocked_without_confirmation_phrase(self, *_m):
        batch = self._make_batch()
        self._login_admin()
        self.client.post(reverse('payroll:archive_clear', args=[batch.pk]),
                         {'confirm_text': 'wrong'})
        self.assertTrue(PayrollRecord.objects.filter(pk=self.payroll.pk).exists())
        batch.refresh_from_db()
        self.assertFalse(batch.is_cleared)

    def test_clear_deletes_only_in_range_records(self, *_m):
        batch = self._make_batch()
        self._login_admin()
        self.client.post(reverse('payroll:archive_clear', args=[batch.pk]),
                         {'confirm_text': 'DELETE ARCHIVED DATA'})
        batch.refresh_from_db()
        self.assertTrue(batch.is_cleared)
        # In-range deleted
        self.assertFalse(PayrollRecord.objects.filter(pk=self.payroll.pk).exists())
        self.assertFalse(AttendanceRecord.objects.filter(pk=self.att.pk).exists())
        self.assertFalse(AttendancePortalLog.objects.filter(pk=self.portal_log.pk).exists())
        self.assertFalse(OvertimeRequest.objects.filter(pk=self.ot.pk).exists())
        self.assertFalse(LeaveRequest.objects.filter(pk=self.leave.pk).exists())
        self.assertFalse(CashAdvanceRequest.objects.filter(pk=self.ca.pk).exists())
        # Out-of-range and other-company survive
        self.assertTrue(AttendanceRecord.objects.filter(pk=self.att_out.pk).exists())
        self.assertTrue(AttendanceRecord.objects.filter(pk=self.other_att.pk).exists())

    def test_clear_preserves_employees_companies_settings(self, *_m):
        batch = self._make_batch()
        self._login_admin()
        self.client.post(reverse('payroll:archive_clear', args=[batch.pk]),
                         {'confirm_text': 'DELETE ARCHIVED DATA'})
        self.assertTrue(Employee.objects.filter(pk=self.emp.pk).exists())
        self.assertTrue(Company.objects.filter(pk=self.company.pk).exists())
        self.assertTrue(AttendanceLocation.objects.filter(pk=self.location.pk).exists())
        self.assertTrue(PayrollPeriod.objects.filter(pk=self.period.pk).exists())
        self.assertTrue(LeaveType.objects.filter(pk=self.ltype.pk).exists())

    def test_clear_blocked_when_range_includes_today(self, *_m):
        # date_to in the future → active-period guard blocks the clear.
        future_batch = ArchiveBatch.objects.create(
            company=self.company, date_from=_FROM,
            date_to=datetime.date.today() + datetime.timedelta(days=5),
            file_name='f.xlsx', file_path='payroll_archives/f.xlsx',
        )
        self._login_admin()
        self.client.post(reverse('payroll:archive_clear', args=[future_batch.pk]),
                         {'confirm_text': 'DELETE ARCHIVED DATA'})
        future_batch.refresh_from_db()
        self.assertFalse(future_batch.is_cleared)
        self.assertTrue(PayrollRecord.objects.filter(pk=self.payroll.pk).exists())

    # ── Permissions / scoping ────────────────────────────────────────────────

    def test_unauthorized_employee_cannot_access(self, *_m):
        user = User.objects.create_user('emp_user', password='pass')
        UserProfile.objects.create(user=user, role='employee', is_active_stafforyx=True)
        self.client.login(username='emp_user', password='pass')
        response = self.client.get(self.url)
        self.assertIn(response.status_code, (302, 403))

    def test_scoped_manager_cannot_download_other_company_batch(self, *_m):
        other_batch = ArchiveBatch.objects.create(
            company=self.other, date_from=_FROM, date_to=_TO,
            file_name='o.xlsx', file_path='payroll_archives/o.xlsx',
        )
        manager = User.objects.create_user('scoped_mgr', password='pass')
        UserProfile.objects.create(
            user=manager, role='hr_admin', is_active_stafforyx=True, can_manage_payroll=True,
        )
        UserCompanyAccess.objects.create(
            user=manager, company=self.company, role='hr_admin', is_active=True,
        )
        self.client.login(username='scoped_mgr', password='pass')
        response = self.client.get(reverse('payroll:archive_download', args=[other_batch.pk]))
        self.assertEqual(response.status_code, 404)


class ArchiveServiceTests(TestCase):
    def test_count_helper_handles_all_sections(self):
        company = Company.objects.create(name='Svc Co', email='svc@test.com')
        qs = collect_archive_querysets(company, _FROM, _TO)
        counts = count_archive_records(qs)
        self.assertEqual(set(counts.keys()), {
            'payroll_records', 'attendance_records', 'portal_logs', 'qr_logs',
            'overtime_requests', 'leave_requests', 'ca_requests',
        })
        self.assertTrue(all(v == 0 for v in counts.values()))


class VacuumSqliteCommandTests(TransactionTestCase):
    def test_vacuum_runs_on_sqlite(self):
        out = StringIO()
        call_command('vacuum_sqlite', stdout=out)
        self.assertIn('VACUUM complete', out.getvalue())

    def test_vacuum_skips_non_sqlite(self):
        out = StringIO()
        fake_conn = MagicMock()
        fake_conn.vendor = 'postgresql'
        with patch('payroll.management.commands.vacuum_sqlite.connection', fake_conn):
            call_command('vacuum_sqlite', stdout=out)
        self.assertIn('skipped', out.getvalue().lower())
        fake_conn.cursor.assert_not_called()


class CleanupCommandTests(TestCase):
    def test_cleanup_requires_existing_batch(self):
        from django.core.management.base import CommandError
        with self.assertRaises(CommandError):
            call_command('cleanup_archived_data', '--archive-batch-id', '99999')
