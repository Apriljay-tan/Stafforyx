import csv
from io import BytesIO

from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone

from announcements.models import Announcement
from attendance.models import AttendanceRecord
from companies.models import Company
from documents.models import EmployeeDocument
from employees.models import Department, Employee
from leaves.models import LeaveRequest, LeaveType
from payroll.models import PayrollPeriod, PayrollRecord


def _csv_response(filename):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _write_csv(filename, headers, rows):
    response = _csv_response(filename)
    writer = csv.writer(response)
    writer.writerow(headers)
    writer.writerows(rows)
    return response


def _employee_queryset(request):
    employees = Employee.objects.select_related('company', 'department', 'position')

    company_id = request.GET.get('company', '')
    if company_id:
        employees = employees.filter(company_id=company_id)

    status = request.GET.get('status', '')
    if status:
        employees = employees.filter(status=status)

    department_id = request.GET.get('department', '')
    if department_id:
        employees = employees.filter(department_id=department_id)

    return employees, {
        'company_filter': company_id,
        'status_filter': status,
        'department_filter': department_id,
    }


def _attendance_queryset(request):
    records = AttendanceRecord.objects.select_related('company', 'employee', 'employee__department')

    employee_id = request.GET.get('employee', '')
    if employee_id:
        records = records.filter(employee_id=employee_id)

    status = request.GET.get('status', '')
    if status:
        records = records.filter(status=status)

    date_from = request.GET.get('date_from', '')
    if date_from:
        records = records.filter(date__gte=date_from)

    date_to = request.GET.get('date_to', '')
    if date_to:
        records = records.filter(date__lte=date_to)

    return records, {
        'employee_filter': employee_id,
        'status_filter': status,
        'date_from_filter': date_from,
        'date_to_filter': date_to,
    }


def _leave_queryset(request):
    leave_requests = LeaveRequest.objects.select_related('company', 'employee', 'leave_type')

    employee_id = request.GET.get('employee', '')
    if employee_id:
        leave_requests = leave_requests.filter(employee_id=employee_id)

    status = request.GET.get('status', '')
    if status:
        leave_requests = leave_requests.filter(status=status)

    leave_type_id = request.GET.get('leave_type', '')
    if leave_type_id:
        leave_requests = leave_requests.filter(leave_type_id=leave_type_id)

    date_from = request.GET.get('date_from', '')
    if date_from:
        leave_requests = leave_requests.filter(start_date__gte=date_from)

    date_to = request.GET.get('date_to', '')
    if date_to:
        leave_requests = leave_requests.filter(end_date__lte=date_to)

    return leave_requests, {
        'employee_filter': employee_id,
        'status_filter': status,
        'leave_type_filter': leave_type_id,
        'date_from_filter': date_from,
        'date_to_filter': date_to,
    }


def _payroll_queryset(request):
    records = PayrollRecord.objects.select_related('company', 'payroll_period', 'employee')

    payroll_period_id = request.GET.get('payroll_period', '')
    if payroll_period_id:
        records = records.filter(payroll_period_id=payroll_period_id)

    employee_id = request.GET.get('employee', '')
    if employee_id:
        records = records.filter(employee_id=employee_id)

    status = request.GET.get('status', '')
    if status:
        records = records.filter(status=status)

    company_id = request.GET.get('company', '')
    if company_id:
        records = records.filter(company_id=company_id)

    return records, {
        'payroll_period_filter': payroll_period_id,
        'employee_filter': employee_id,
        'status_filter': status,
        'company_filter': company_id,
    }


def _documents_queryset(request):
    documents = EmployeeDocument.objects.select_related('company', 'employee', 'employee__department')

    employee_id = request.GET.get('employee', '')
    if employee_id:
        documents = documents.filter(employee_id=employee_id)

    document_type = request.GET.get('document_type', '')
    if document_type:
        documents = documents.filter(document_type=document_type)

    company_id = request.GET.get('company', '')
    if company_id:
        documents = documents.filter(company_id=company_id)

    return documents, {
        'employee_filter': employee_id,
        'document_type_filter': document_type,
        'company_filter': company_id,
    }


def reports_dashboard(request):
    context = {
        'employee_count': Employee.objects.count(),
        'attendance_count': AttendanceRecord.objects.count(),
        'leave_count': LeaveRequest.objects.count(),
        'payroll_count': PayrollRecord.objects.count(),
        'document_count': EmployeeDocument.objects.count(),
    }
    return render(request, 'reports/reports_dashboard.html', context)


def _apply_excel_layout(ws, title, export_time, currency_columns=None, date_columns=None):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    currency_columns = currency_columns or []
    date_columns = date_columns or []

    max_column = max(ws.max_column, 1)
    title_fill = PatternFill('solid', fgColor='0D1B2A')
    header_fill = PatternFill('solid', fgColor='1565C0')
    thin_side = Side(style='thin', color='D9E2EC')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    ws.insert_rows(1, 3)
    ws.cell(row=1, column=1, value='Stafforyx HR')
    ws.cell(row=2, column=1, value=f'{title} - Exported {export_time:%Y-%m-%d %I:%M %p}')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_column)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_column)

    ws.cell(row=1, column=1).font = Font(color='FFFFFF', bold=True, size=16)
    ws.cell(row=1, column=1).fill = title_fill
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    ws.cell(row=2, column=1).font = Font(color='334155', italic=True)

    header_row = 4
    ws.freeze_panes = 'A5'

    for cell in ws[header_row]:
        cell.font = Font(color='FFFFFF', bold=True)
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    for column_index in currency_columns:
        column_letter = get_column_letter(column_index)
        for cell in ws[column_letter][4:]:
            cell.number_format = '"PHP" #,##0.00'

    for column_index in date_columns:
        column_letter = get_column_letter(column_index)
        for cell in ws[column_letter][4:]:
            cell.number_format = 'yyyy-mm-dd'

    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 42)


def _add_sheet(wb, title, headers, rows, export_time, currency_columns=None, date_columns=None):
    ws = wb.create_sheet(title=title)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    _apply_excel_layout(ws, title, export_time, currency_columns, date_columns)
    return ws


def export_all_data(request):
    from openpyxl import Workbook

    export_time = timezone.localtime()
    wb = Workbook()
    wb.remove(wb.active)

    summary_rows = [
        ['Total Companies', Company.objects.count()],
        ['Total Active Employees', Employee.objects.filter(status='active').count()],
        ['Total Attendance Records', AttendanceRecord.objects.count()],
        ['Total Leave Requests', LeaveRequest.objects.count()],
        ['Total Payroll Records', PayrollRecord.objects.count()],
        ['Total Documents', EmployeeDocument.objects.count()],
        ['Total Announcements', Announcement.objects.count()],
    ]
    _add_sheet(
        wb,
        'Summary',
        ['Metric', 'Value'],
        summary_rows,
        export_time,
    )

    _add_sheet(
        wb,
        'Companies',
        ['Name', 'Email', 'Phone', 'Status', 'Plan', 'Created At'],
        [
            [
                company.name,
                company.email,
                company.phone,
                company.get_status_display(),
                company.get_subscription_plan_display(),
                timezone.localtime(company.created_at).date(),
            ]
            for company in Company.objects.all()
        ],
        export_time,
        date_columns=[6],
    )

    _add_sheet(
        wb,
        'Employees',
        ['Employee ID', 'Name', 'Email', 'Company', 'Department', 'Position', 'Type', 'Status', 'Date Hired', 'Basic Salary'],
        [
            [
                employee.employee_id,
                employee.full_name,
                employee.email,
                employee.company.name,
                employee.department.name if employee.department else '',
                employee.position.title if employee.position else '',
                employee.get_employment_type_display(),
                employee.get_status_display(),
                employee.date_hired,
                employee.basic_salary,
            ]
            for employee in Employee.objects.select_related('company', 'department', 'position')
        ],
        export_time,
        currency_columns=[10],
        date_columns=[9],
    )

    _add_sheet(
        wb,
        'Attendance',
        [
            'Date', 'Employee ID', 'Employee', 'Company',
            'Time In', 'Time Out', 'Break Min',
            'Total Hours', 'Total Work Min',
            'Late Min', 'Undertime Min', 'Overtime Min',
            'Status', 'Computed Status',
        ],
        [
            [
                record.date,
                record.employee.employee_id,
                record.employee.full_name,
                record.company.name,
                record.time_in,
                record.time_out,
                record.break_minutes,
                record.total_hours,
                record.total_work_minutes,
                record.late_minutes,
                record.undertime_minutes,
                record.overtime_minutes,
                record.get_status_display(),
                record.computed_status,
            ]
            for record in AttendanceRecord.objects.select_related('company', 'employee')
        ],
        export_time,
        date_columns=[1],
    )

    _add_sheet(
        wb,
        'Leave Requests',
        ['Employee ID', 'Employee', 'Company', 'Leave Type', 'Start Date', 'End Date', 'Total Days', 'Status', 'Reviewed At', 'Reason', 'Remarks'],
        [
            [
                leave_request.employee.employee_id,
                leave_request.employee.full_name,
                leave_request.company.name,
                leave_request.leave_type.name,
                leave_request.start_date,
                leave_request.end_date,
                leave_request.total_days,
                leave_request.get_status_display(),
                timezone.localtime(leave_request.reviewed_at).date() if leave_request.reviewed_at else '',
                leave_request.reason,
                leave_request.remarks,
            ]
            for leave_request in LeaveRequest.objects.select_related('company', 'employee', 'leave_type')
        ],
        export_time,
        date_columns=[5, 6, 9],
    )

    _add_sheet(
        wb,
        'Payroll Records',
        [
            'Payroll Period', 'Employee ID', 'Employee', 'Company', 'Basic Pay',
            'Allowances', 'Overtime Pay', 'Gross Pay', 'SSS Deduction',
            'PhilHealth Deduction', 'Pag-IBIG Deduction', 'Tax Deduction',
            'Late Deduction', 'Absence Deduction', 'Other Deductions', 'Net Pay', 'Status',
        ],
        [
            [
                record.payroll_period.name,
                record.employee.employee_id,
                record.employee.full_name,
                record.company.name,
                record.basic_pay,
                record.allowances,
                record.overtime_pay,
                record.gross_pay,
                record.sss_deduction,
                record.philhealth_deduction,
                record.pagibig_deduction,
                record.tax_deduction,
                record.late_deduction,
                record.absence_deduction,
                record.other_deductions,
                record.net_pay,
                record.get_status_display(),
            ]
            for record in PayrollRecord.objects.select_related('company', 'payroll_period', 'employee')
        ],
        export_time,
        currency_columns=[5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16],
    )

    _add_sheet(
        wb,
        'Employee Documents',
        ['Employee ID', 'Employee', 'Company', 'Title', 'Type', 'Expiration Date', 'File', 'Notes'],
        [
            [
                document.employee.employee_id,
                document.employee.full_name,
                document.company.name,
                document.title,
                document.get_document_type_display(),
                document.expiration_date,
                document.file.url if document.file else '',
                document.notes,
            ]
            for document in EmployeeDocument.objects.select_related('company', 'employee')
        ],
        export_time,
        date_columns=[6],
    )

    _add_sheet(
        wb,
        'Announcements',
        ['Title', 'Company', 'Target Department', 'Active', 'Posted By', 'Created At', 'Content'],
        [
            [
                announcement.title,
                announcement.company.name,
                announcement.target_department.name if announcement.target_department else 'All Departments',
                'Active' if announcement.is_active else 'Inactive',
                announcement.posted_by.get_full_name() or announcement.posted_by.username if announcement.posted_by else '',
                timezone.localtime(announcement.created_at).date(),
                announcement.content,
            ]
            for announcement in Announcement.objects.select_related('company', 'target_department', 'posted_by')
        ],
        export_time,
        date_columns=[6],
    )

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="stafforyx_all_data_export.xlsx"'
    return response


def employee_report(request):
    employees, filters = _employee_queryset(request)
    context = {
        'employees': employees,
        'companies': Company.objects.all(),
        'departments': Department.objects.select_related('company'),
        'status_choices': Employee.STATUS_CHOICES,
        **filters,
    }
    return render(request, 'reports/employee_report.html', context)


def employee_report_export(request):
    employees, _filters = _employee_queryset(request)
    rows = [
        [
            employee.employee_id,
            employee.full_name,
            employee.email,
            employee.company.name,
            employee.department.name if employee.department else '',
            employee.position.title if employee.position else '',
            employee.get_status_display(),
            employee.date_hired,
            employee.basic_salary,
        ]
        for employee in employees
    ]
    return _write_csv(
        'employee_report.csv',
        ['Employee ID', 'Name', 'Email', 'Company', 'Department', 'Position', 'Status', 'Date Hired', 'Basic Salary'],
        rows,
    )


def attendance_report(request):
    records, filters = _attendance_queryset(request)
    context = {
        'records': records,
        'employees': Employee.objects.all().order_by('last_name', 'first_name'),
        'status_choices': AttendanceRecord.STATUS_CHOICES,
        **filters,
    }
    return render(request, 'reports/attendance_report.html', context)


def attendance_report_export(request):
    records, _filters = _attendance_queryset(request)
    rows = [
        [
            record.date,
            record.employee.employee_id,
            record.employee.full_name,
            record.time_in or '',
            record.time_out or '',
            record.total_hours,
            record.late_minutes,
            record.overtime_hours,
            record.get_status_display(),
        ]
        for record in records
    ]
    return _write_csv(
        'attendance_report.csv',
        ['Date', 'Employee ID', 'Employee', 'Time In', 'Time Out', 'Total Hours', 'Late Minutes', 'Overtime Hours', 'Status'],
        rows,
    )


def leave_report(request):
    leave_requests, filters = _leave_queryset(request)
    context = {
        'leave_requests': leave_requests,
        'employees': Employee.objects.all().order_by('last_name', 'first_name'),
        'leave_types': LeaveType.objects.select_related('company'),
        'status_choices': LeaveRequest.STATUS_CHOICES,
        **filters,
    }
    return render(request, 'reports/leave_report.html', context)


def leave_report_export(request):
    leave_requests, _filters = _leave_queryset(request)
    rows = [
        [
            leave_request.employee.employee_id,
            leave_request.employee.full_name,
            leave_request.leave_type.name,
            leave_request.start_date,
            leave_request.end_date,
            leave_request.total_days,
            leave_request.get_status_display(),
            leave_request.reason,
        ]
        for leave_request in leave_requests
    ]
    return _write_csv(
        'leave_report.csv',
        ['Employee ID', 'Employee', 'Leave Type', 'Start Date', 'End Date', 'Total Days', 'Status', 'Reason'],
        rows,
    )


def payroll_report(request):
    records, filters = _payroll_queryset(request)
    context = {
        'records': records,
        'payroll_periods': PayrollPeriod.objects.select_related('company'),
        'employees': Employee.objects.all().order_by('last_name', 'first_name'),
        'companies': Company.objects.all(),
        'status_choices': PayrollRecord.STATUS_CHOICES,
        **filters,
    }
    return render(request, 'reports/payroll_report.html', context)


def payroll_report_export(request):
    records, _filters = _payroll_queryset(request)
    rows = [
        [
            record.payroll_period.name,
            record.employee.employee_id,
            record.employee.full_name,
            record.company.name,
            record.basic_pay,
            record.allowances,
            record.overtime_pay,
            record.gross_pay,
            record.sss_deduction,
            record.philhealth_deduction,
            record.pagibig_deduction,
            record.tax_deduction,
            record.late_deduction,
            record.absence_deduction,
            record.other_deductions,
            record.net_pay,
            record.get_status_display(),
        ]
        for record in records
    ]
    return _write_csv(
        'payroll_report.csv',
        [
            'Payroll Period', 'Employee ID', 'Employee', 'Company', 'Basic Pay',
            'Allowances', 'Overtime Pay', 'Gross Pay', 'SSS Deduction',
            'PhilHealth Deduction', 'Pag-IBIG Deduction', 'Tax Deduction',
            'Late Deduction', 'Absence Deduction', 'Other Deductions', 'Net Pay', 'Status',
        ],
        rows,
    )


def documents_report(request):
    documents, filters = _documents_queryset(request)
    context = {
        'documents': documents,
        'employees': Employee.objects.all().order_by('last_name', 'first_name'),
        'companies': Company.objects.all(),
        'document_type_choices': EmployeeDocument.DOCUMENT_TYPE_CHOICES,
        **filters,
    }
    return render(request, 'reports/documents_report.html', context)


def documents_report_export(request):
    documents, _filters = _documents_queryset(request)
    rows = [
        [
            document.employee.employee_id,
            document.employee.full_name,
            document.company.name,
            document.title,
            document.get_document_type_display(),
            document.expiration_date or '',
            document.file.url if document.file else '',
            document.notes,
        ]
        for document in documents
    ]
    return _write_csv(
        'documents_report.csv',
        ['Employee ID', 'Employee', 'Company', 'Title', 'Document Type', 'Expiration Date', 'File URL', 'Notes'],
        rows,
    )
